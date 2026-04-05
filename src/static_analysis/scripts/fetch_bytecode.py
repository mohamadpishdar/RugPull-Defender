#!/usr/bin/env python3
"""
fetch_bytecode.py - Collect runtime bytecode from BSC and ETH blockchains for contracts in dataset.
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('fetch_bytecode.log')
    ]
)
logger = logging.getLogger(__name__)

# Chain configurations
CHAIN_CONFIGS = {
    'BSC': { # end point has deleted
        'rpc_url': 'https://go.getblock.io/a31d3b5f43b2461ca12fdfb57303b665',
        'explorer_domain': 'bscscan.com',
        'explorer_api_url': 'https://api.bscscan.com/api',
        'chain_name': 'bsc'
    },
    'ETH': {
        'rpc_url': 'https://go.getblock.io/e85b270e84284f4a9fd42b6d55556b73',
        'explorer_domain': 'etherscan.io',
        'explorer_api_url': 'https://api.etherscan.io/api',
        'chain_name': 'eth'
    },
    'ARBI': {
        'rpc_url': 'https://arb-mainnet.g.alchemy.com/v2/InnLbi27UJ9Ge3sseHHqOVjncP1Q5VnY',
        'explorer_domain': 'arbiscan.io',
        'explorer_api_url': 'https://api.arbiscan.io/api',
        'chain_name': 'arbi'
    },
    'POLYGON': {
        'rpc_url': 'https://polygon-mainnet.g.alchemy.com/v2/Nw1B39O8HWajcz1P0ZtZoCvWAeKsA1yM',
        'explorer_domain': 'polygonscan.com',
        'explorer_api_url': 'https://api.polygonscan.com/api',
        'chain_name': 'polygon'
    },
    'FANTOM': {
        'rpc_url': 'https://fantom-mainnet.core.chainstack.com/d9a0800044449a8f6e0e91032c64a3cd',
        'explorer_domain': 'ftmscan.com',
        'explorer_api_url': 'https://api.ftmscan.com/api',
        'chain_name': 'fantom'
    },
    'CRONO': {
        'rpc_url': 'https://go.getblock.io/06c4d0ce64a74cfda34714871d286a91',
        'explorer_domain': 'explorer.cronos.org',
        'explorer_api_url': 'https://api.cronoscan.com/api',
        'chain_name': 'crono'
    },
    'BASE': {
        'rpc_url': 'https://base-mainnet.g.alchemy.com/v2/InnLbi27UJ9Ge3sseHHqOVjncP1Q5VnY',
        'explorer_domain': 'basescan.org',
        'explorer_api_url': 'https://api.basescan.org/api',
        'chain_name': 'base'
    }
}


class BytecodeFetcher:
    """Fetches bytecode from blockchain."""
    
    def __init__(self, chain_type: str, api_key: Optional[str] = None):
        self.chain_type = chain_type.upper()
        if self.chain_type not in CHAIN_CONFIGS:
            raise ValueError(f"Unsupported chain type: {chain_type}")
        
        config = CHAIN_CONFIGS[self.chain_type]
        self.rpc_url = config['rpc_url']
        self.explorer_domain = config['explorer_domain']
        self.explorer_api_url = config['explorer_api_url']
        self.chain_name = config['chain_name']
        self.api_key = api_key or ""
        
        # Setup session with retries
        self.session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
        
        logger.info(f"Initialized fetcher for {self.chain_type} chain")
        
    def _make_rpc_call(self, method: str, params: list) -> Optional[Dict]:
        """Make JSON-RPC call to blockchain node."""
        payload = {
            "jsonrpc": "2.0",
            "method": method,
            "params": params,
            "id": 1
        }
        try:
            response = self.session.post(
                self.rpc_url,
                json=payload,
                timeout=30,
                headers={"Content-Type": "application/json"}
            )
            response.raise_for_status()
            data = response.json()
            if "error" in data:
                logger.error(f"RPC error: {data['error']}")
                return None
            return data.get("result")
        except Exception as e:
            logger.error(f"RPC call failed: {e}")
            return None
    
    def _make_explorer_call(self, params: Dict) -> Optional[Dict]:
        """Make API call to blockchain explorer."""
        params["apikey"] = self.api_key
        try:
            response = self.session.get(
                self.explorer_api_url,
                params=params,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            if data.get("status") == "1":
                return data.get("result")
            else:
                logger.warning(f"Explorer API error: {data.get('message', 'Unknown error')}")
                return None
        except Exception as e:
            logger.error(f"Explorer API call failed: {e}")
            return None
    
    def extract_address_from_url(self, url: str) -> Tuple[Optional[str], str]:
        """
        Extract contract address from various URL types.
        Returns (address, resolved_from_type)
        """
        if not url or pd.isna(url):
            return None, "none"
        
        url = url.strip()
        
        # Pattern 1: Direct address page (address/0x...)
        address_pattern = rf'{self.explorer_domain}/address/(0x[a-fA-F0-9]{{40}})'
        match = re.search(address_pattern, url, re.IGNORECASE)
        if match:
            return match.group(1), "address"
        
        # Pattern 2: Token page (token/0x...)
        token_pattern = rf'{self.explorer_domain}/token/(0x[a-fA-F0-9]{{40}})'
        match = re.search(token_pattern, url, re.IGNORECASE)
        if match:
            address = match.group(1)
            logger.info(f"Found token page, extracting address: {address}")
            return address, "token"
        
        # Pattern 3: External website - fetch and search for explorer link
        if self.explorer_domain not in url.lower():
            logger.info(f"External URL detected, fetching page: {url}")
            try:
                response = self.session.get(url, timeout=15, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                response.raise_for_status()
                html = response.text
                
                # Search for explorer address links in HTML
                explorer_links = re.findall(
                    rf'https?://{self.explorer_domain}/(?:address|token)/(0x[a-fA-F0-9]{{40}})',
                    html,
                    re.IGNORECASE
                )
                if explorer_links:
                    address = explorer_links[0]
                    logger.info(f"Found {self.explorer_domain} address in external page: {address}")
                    return address, "external_page"
            except Exception as e:
                logger.error(f"Failed to fetch external page {url}: {e}")
        
        return None, "none"
    
    def fetch_runtime_bytecode(self, address: str) -> Tuple[Optional[str], str, Optional[str]]:
        """
        Fetch runtime bytecode for given address.
        Returns (bytecode, fetch_method, block_number)
        """
        # Try RPC first
        logger.info(f"Fetching bytecode via RPC for {address}")
        result = self._make_rpc_call("eth_getCode", [address, "latest"])
        if result and result != "0x":
            block_result = self._make_rpc_call("eth_blockNumber", [])
            return result, "rpc", block_result
        
        # Fallback to Explorer proxy API
        logger.info(f"RPC failed or empty, trying {self.explorer_domain} proxy for {address}")
        params = {
            "module": "proxy",
            "action": "eth_getCode",
            "address": address,
            "tag": "latest"
        }
        result = self._make_explorer_call(params)
        if result and result != "0x":
            # Get current block via Explorer
            block_params = {
                "module": "proxy",
                "action": "eth_blockNumber"
            }
            block_result = self._make_explorer_call(block_params)
            return result, f"{self.chain_name}_proxy", block_result
        
        return None, "failed", None
    
    def fetch_creation_info(self, address: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Fetch creation transaction and creation bytecode.
        Returns (creation_tx_hash, creation_bytecode)
        """
        logger.info(f"Fetching creation info for {address}")
        
        # Get creation transaction
        params = {
            "module": "contract",
            "action": "getcontractcreation",
            "contractaddresses": address
        }
        result = self._make_explorer_call(params)
        
        if not result or not isinstance(result, list) or len(result) == 0:
            logger.warning(f"No creation info found for {address}")
            return None, None
        
        creation_tx = result[0].get("txHash")
        if not creation_tx:
            return None, None
        
        logger.info(f"Found creation tx: {creation_tx}")
        
        # Fetch transaction details to get input data (creation bytecode)
        tx_params = {
            "module": "proxy",
            "action": "eth_getTransactionByHash",
            "txhash": creation_tx
        }
        tx_result = self._make_explorer_call(tx_params)
        
        if tx_result and "input" in tx_result:
            creation_bytecode = tx_result["input"]
            logger.info(f"Retrieved creation bytecode ({len(creation_bytecode)} chars)")
            return creation_tx, creation_bytecode
        
        return creation_tx, None


def extract_hyperlinks_from_excel(file_path: str, column_name: str) -> Dict[int, str]:
    """
    Extract hyperlinks from Excel file using openpyxl.
    Returns dictionary mapping row index to URL.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return {}
    
    file_path = Path(file_path)
    if file_path.suffix.lower() not in ['.xlsx', '.xls']:
        return {}
    
    logger.info(f"Extracting hyperlinks from Excel file: {file_path}")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find column index for the specified column name
        headers = [cell.value for cell in ws[1]]
        if column_name not in headers:
            logger.warning(f"Column '{column_name}' not found in Excel file")
            return {}
        
        col_idx = headers.index(column_name) + 1  # openpyxl uses 1-based indexing
        
        hyperlinks = {}
        for row_idx, cell in enumerate(ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx), start=0):
            cell = cell[0]
            if cell.hyperlink:
                url = cell.hyperlink.target
                hyperlinks[row_idx] = url
                logger.debug(f"Row {row_idx}: Found hyperlink: {url}")
            elif cell.value and isinstance(cell.value, str):
                # Check if the cell value itself is a URL
                if cell.value.startswith('http://') or cell.value.startswith('https://'):
                    hyperlinks[row_idx] = cell.value
                    logger.debug(f"Row {row_idx}: Found URL in cell value: {cell.value}")
        
        logger.info(f"Extracted {len(hyperlinks)} hyperlinks from column '{column_name}'")
        return hyperlinks
        
    except Exception as e:
        logger.error(f"Failed to extract hyperlinks from Excel: {e}")
        return {}


def load_dataset(file_path: str) -> pd.DataFrame:
    """Load dataset from CSV or XLSX file."""
    file_path_obj = Path(file_path)
    
    if not file_path_obj.exists():
        raise FileNotFoundError(f"Dataset file not found: {file_path_obj}")
    
    if file_path_obj.suffix.lower() == '.csv':
        df = pd.read_csv(file_path)
        logger.info(f"Loaded CSV dataset with {len(df)} rows")
    elif file_path_obj.suffix.lower() in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path)
        logger.info(f"Loaded Excel dataset with {len(df)} rows")
        
        # Extract hyperlinks from 'Smart Contract Online' column
        hyperlinks = extract_hyperlinks_from_excel(file_path, 'Smart Contract Online')
        
        if hyperlinks:
            # Create a new column with actual URLs
            df['_extracted_url'] = df.index.map(lambda idx: hyperlinks.get(idx))
            
            # Log comparison
            for idx in range(min(5, len(df))):
                display_text = df.loc[idx, 'Smart Contract Online']
                actual_url = df.loc[idx, '_extracted_url']
                if actual_url:
                    logger.info(f"Row {idx}: '{display_text}' -> '{actual_url}'")
        else:
            logger.warning("No hyperlinks extracted from Excel file. Using cell values as-is.")
            df['_extracted_url'] = df['Smart Contract Online']
    else:
        raise ValueError(f"Unsupported file format: {file_path_obj.suffix}")
    
    return df


def save_bytecode_data(
    output_dir: Path,
    contract_id: str,
    runtime_bytecode: str,
    creation_bytecode: Optional[str],
    metadata: Dict
):
    """Save bytecode and metadata to files."""
    contract_dir = output_dir / str(contract_id)
    contract_dir.mkdir(parents=True, exist_ok=True)
    
    # Save runtime bytecode
    runtime_file = contract_dir / f"{contract_id}.hex"
    with open(runtime_file, 'w') as f:
        f.write(runtime_bytecode)
    logger.info(f"Saved runtime bytecode: {runtime_file}")
    
    # Save creation bytecode if available
    if creation_bytecode:
        creation_file = contract_dir / f"{contract_id}.creation.hex"
        with open(creation_file, 'w') as f:
            f.write(creation_bytecode)
        logger.info(f"Saved creation bytecode: {creation_file}")
    
    # Save metadata
    meta_file = contract_dir / f"{contract_id}.meta.json"
    with open(meta_file, 'w') as f:
        json.dump(metadata, f, indent=2)
    logger.info(f"Saved metadata: {meta_file}")


def process_dataset(
    dataset_path: str,
    output_dir: str,
    chain_types: list,
    api_keys: Dict[str, str] = None,
    delay: float = 0.5
):
    """Process entire dataset and fetch bytecode for all specified blockchain contracts."""
    df = load_dataset(dataset_path)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    if api_keys is None:
        api_keys = {}
    
    # Use extracted URL column if available, otherwise use original column
    url_column = '_extracted_url' if '_extracted_url' in df.columns else 'Smart Contract Online'
    
    overall_stats = {
        "total": 0,
        "success": 0,
        "failed": 0,
        "skipped": 0
    }
    
    for chain_type in chain_types:
        chain_type_upper = chain_type.upper()
        logger.info(f"\n{'#'*80}")
        logger.info(f"# PROCESSING {chain_type_upper} CONTRACTS")
        logger.info(f"{'#'*80}\n")
        
        # Filter for specific blockchain contracts
        chain_contracts = df[df['Blockchain Type'].str.upper() == chain_type_upper].copy()
        logger.info(f"Found {len(chain_contracts)} {chain_type_upper} contracts to process")
        
        if len(chain_contracts) == 0:
            logger.warning(f"No {chain_type_upper} contracts found in dataset")
            continue
        
        # Initialize fetcher for this chain
        api_key = api_keys.get(chain_type_upper, None)
        fetcher = BytecodeFetcher(chain_type_upper, api_key)
        
        chain_stats = {
            "total": len(chain_contracts),
            "success": 0,
            "failed": 0,
            "skipped": 0
        }
        
        for idx, row in chain_contracts.iterrows():
            contract_id = str(row['Smart Contract Offline'])
            title = row['title']
            source_url = row[url_column]
            display_text = row['Smart Contract Online']
            
            logger.info(f"\n{'='*80}")
            logger.info(f"Processing [{idx+1}/{len(df)}] {chain_type_upper} ID: {contract_id} - {title}")
            if url_column == '_extracted_url':
                logger.info(f"Display text: {display_text}")
                logger.info(f"Actual URL: {source_url}")
            else:
                logger.info(f"URL: {source_url}")
            logger.info(f"{'='*80}")
            
            errors = []
            
            # Check if already processed
            contract_dir = output_path / contract_id
            if contract_dir.exists() and (contract_dir / f"{contract_id}.hex").exists():
                logger.info(f"Skipping {contract_id} - already processed")
                chain_stats["skipped"] += 1
                continue
            
            # Step 1: Resolve contract address
            address, resolved_from = fetcher.extract_address_from_url(source_url)
            if not address:
                error_msg = f"Could not resolve contract address from URL: {source_url}"
                logger.error(error_msg)
                errors.append(error_msg)
                chain_stats["failed"] += 1
                
                # Save metadata with error
                metadata = {
                    "title": title,
                    "chain": fetcher.chain_name,
                    "address": None,
                    "source_url": source_url,
                    "display_text": display_text,
                    "resolved_from": resolved_from,
                    "errors": errors
                }
                contract_dir.mkdir(parents=True, exist_ok=True)
                meta_file = contract_dir / f"{contract_id}.meta.json"
                with open(meta_file, 'w') as f:
                    json.dump(metadata, f, indent=2)
                
                time.sleep(delay)
                continue
            
            logger.info(f"Resolved address: {address} (from: {resolved_from})")
            
            # Step 2: Fetch runtime bytecode
            runtime_bytecode, fetch_method, block_number = fetcher.fetch_runtime_bytecode(address)
            if not runtime_bytecode or runtime_bytecode == "0x":
                error_msg = f"Failed to fetch runtime bytecode for {address}"
                logger.error(error_msg)
                errors.append(error_msg)
                chain_stats["failed"] += 1
                
                # Save metadata with error
                metadata = {
                    "title": title,
                    "chain": fetcher.chain_name,
                    "address": address,
                    "source_url": source_url,
                    "display_text": display_text,
                    "resolved_from": resolved_from,
                    "fetch_method": fetch_method,
                    "errors": errors
                }
                contract_dir.mkdir(parents=True, exist_ok=True)
                meta_file = contract_dir / f"{contract_id}.meta.json"
                with open(meta_file, 'w') as f:
                    json.dump(metadata, f, indent=2)
                
                time.sleep(delay)
                continue
            
            logger.info(f"Fetched runtime bytecode: {len(runtime_bytecode)} chars via {fetch_method}")
            
            # Step 3: Fetch creation info
            creation_tx, creation_bytecode = fetcher.fetch_creation_info(address)
            if not creation_tx:
                errors.append("Creation transaction not found")
            if not creation_bytecode:
                errors.append("Creation bytecode not available")
            
            # Step 4: Save all data
            metadata = {
                "title": title,
                "chain": fetcher.chain_name,
                "address": address,
                "source_url": source_url,
                "display_text": display_text,
                "resolved_from": resolved_from,
                "fetch_method": fetch_method,
                "fetched_at_block": block_number,
                "creation_tx": creation_tx,
                "errors": errors
            }
            
            save_bytecode_data(
                output_path,
                contract_id,
                runtime_bytecode,
                creation_bytecode,
                metadata
            )
            
            chain_stats["success"] += 1
            logger.info(f"✓ Successfully processed {chain_type_upper} contract {contract_id}")
            
            # Rate limiting
            time.sleep(delay)
        
        # Print chain summary
        logger.info(f"\n{'='*80}")
        logger.info(f"{chain_type_upper} CHAIN SUMMARY")
        logger.info(f"{'='*80}")
        logger.info(f"Total contracts: {chain_stats['total']}")
        logger.info(f"Successfully processed: {chain_stats['success']}")
        logger.info(f"Failed: {chain_stats['failed']}")
        logger.info(f"Skipped (already processed): {chain_stats['skipped']}")
        logger.info(f"{'='*80}\n")
        
        # Update overall stats
        overall_stats["total"] += chain_stats["total"]
        overall_stats["success"] += chain_stats["success"]
        overall_stats["failed"] += chain_stats["failed"]
        overall_stats["skipped"] += chain_stats["skipped"]
    
    # Print overall summary
    logger.info(f"\n{'#'*80}")
    logger.info("# OVERALL PROCESSING COMPLETE")
    logger.info(f"{'#'*80}")
    logger.info(f"Total contracts: {overall_stats['total']}")
    logger.info(f"Successfully processed: {overall_stats['success']}")
    logger.info(f"Failed: {overall_stats['failed']}")
    logger.info(f"Skipped (already processed): {overall_stats['skipped']}")
    logger.info(f"{'#'*80}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Fetch runtime bytecode from BSC and ETH blockchains for contracts in dataset"
    )
    parser.add_argument(
        "dataset",
        help="Path to input CSV or XLSX dataset file"
    )
    parser.add_argument(
        "-o", "--output",
        default="dataset/bytecode",
        help="Output directory for bytecode files (default: dataset/bytecode)"
    )
    parser.add_argument(
        "-c", "--chains",
        nargs="+",
        default=["ETH"],
        choices=[
        "BSC", "ETH", "ARBI", "POLYGON", "FANTOM", "CRONO", "BASE",
        "bsc", "eth", "arbi", "polygon", "fantom", "crono", "base"
    ],
        help="Blockchain types to process (default: ETH)"
    )
    parser.add_argument(
        "--bsc-api-key",
        help="BscScan API key (optional, for higher rate limits)"
    )
    parser.add_argument(
        "--eth-api-key",
        help="Etherscan API key (optional, for higher rate limits)"
    )
    parser.add_argument(
        "--arbi-api-key",
        help="Arbiscan API key (optional, for higher rate limits)"
    )
    parser.add_argument(
        "--polygon-api-key",
        help="Polygonscan API key (optional, for higher rate limits)"
    )
    parser.add_argument(
        "-d", "--delay",
        type=float,
        default=0.5,
        help="Delay between requests in seconds (default: 0.5)"
    )
    
    args = parser.parse_args()
    
    # Prepare API keys dictionary
    api_keys = {}
    if args.bsc_api_key:
        api_keys['BSC'] = args.bsc_api_key
    if args.eth_api_key:
        api_keys['ETH'] = args.eth_api_key
    if args.arbi_api_key:
        api_keys['ARBI'] = args.arbi_api_key
    if args.polygon_api_key:
        api_keys['POLYGON'] = args.polygon_api_key
    
    # Normalize chain types to uppercase
    chain_types = [c.upper() for c in args.chains]
    
    try:
        process_dataset(
            args.dataset,
            args.output,
            chain_types,
            api_keys,
            args.delay
        )
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()