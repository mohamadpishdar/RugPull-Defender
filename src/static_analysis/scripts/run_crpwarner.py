#!/usr/bin/env python3
import re
import subprocess
import sys
import shutil
import time
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ----------------------------
# CONFIG — adjust if needed
# ----------------------------
PROJECT_ROOT = Path("/home/senatoma/coinfa/gigahorse-crpwarner/RugPull/gigahorse").resolve()
CLIENT_DATALOG = PROJECT_ROOT / "clients" / "crpwarner.dl"

# NOTE: Updated to read directly from the new compiled path for .hex files.
COMPILED_DIR = Path("/home/senatoma/coinfa/RugPullHunter/dataset/bytecode").resolve()

# Gigahorse’s temp cache (per CWD)
TEMP_DIR = PROJECT_ROOT / ".temp"

# Outputs
CRPWARNER_RESULTS = Path("/home/senatoma/coinfa/RugPullHunter/dataset/crpwarner_results").resolve()
OUTPUT_DIR = CRPWARNER_RESULTS / "all_bytecodes_batch_outputs"
OUTPUT_XLSX = OUTPUT_DIR / "crpwarner_results_all.xlsx"
LOG_DIR = OUTPUT_DIR / "logs"

PYTHON_BIN = sys.executable or "python3"
ANALYSIS_TIMEOUT = 1200
OVERWRITE_RESULTS = False  # overwrite Excel each run; set False to append across runs

# ----------------------------
# Helpers
# ----------------------------
def ensure_dirs():
    """Ensure output directories exist."""
    for d in (OUTPUT_DIR, LOG_DIR):
        d.mkdir(parents=True, exist_ok=True)

def init_workbook():
    """Initializes or loads the results Excel workbook."""
    if OVERWRITE_RESULTS or not OUTPUT_XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "CRPWarner"
        headers = [
            "GroupID",         # e.g., 2, 3, 4...
            "Artifact",        # The name of the hex file, e.g., token.hex
            "InputHexFile",    # Renamed from StagedHex (now the direct input path)
            "HiddenMint",
            "LeakingToken",
            "LimitingSell",
            "RugPull",
            "Timestamp"
        ]
        ws.append(headers)
        widths = [10, 32, 40, 14, 14, 14, 10, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(OUTPUT_XLSX)
        return wb, ws
    else:
        from openpyxl import load_workbook
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        return wb, ws

def save_workbook(wb):
    """Saves the Excel workbook."""
    wb.save(OUTPUT_XLSX)

def clear_contract_cache(input_hex_path: Path):
    """
    Remove only this input hex's cache: Gigahorse uses the basename (stem) under .temp/.
    """
    # Gigahorse uses the filename stem (e.g., 'token') for its cache folder name
    cache_dir = TEMP_DIR / input_hex_path.stem
    if cache_dir.exists():
        shutil.rmtree(cache_dir)

def run_crpwarner_on(input_hex_path: Path):
    """Runs the CRPWarner client on the provided .hex file."""
    cmd = [
        PYTHON_BIN, str(PROJECT_ROOT / "gigahorse.py"),
        "--debug",
        "-C", str(CLIENT_DATALOG),
        str(input_hex_path)
    ]
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(PROJECT_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=ANALYSIS_TIMEOUT
        )
        return proc.returncode == 0, proc.stdout, proc.stderr
    except subprocess.TimeoutExpired as e:
        return False, "", f"TimeoutExpired: {e}"
    except Exception as e:
        return False, "", f"Exception: {e}"

# Regexes to parse console flags
RE_HIDDEN = re.compile(r"Hidden\s+Mint\s+Function:\s*(True|False)", re.IGNORECASE)
RE_LEAK   = re.compile(r"Leaking\s+Token:\s*(True|False)", re.IGNORECASE)
RE_LIMIT  = re.compile(r"Limiting\s+Sell\s+Order:\s*(True|False)", re.IGNORECASE)

def parse_flags(stdout: str):
    """Parses the boolean flags from the analysis stdout."""
    def b(m): return None if m is None else (m.group(1).strip().lower() == "true")
    return b(RE_HIDDEN.search(stdout)), b(RE_LEAK.search(stdout)), b(RE_LIMIT.search(stdout))

def find_row_by_group_id(ws, group_id: int):
    """
    Finds the row number (1-indexed) where GroupID matches.
    Returns None if not found.
    Assumes GroupID is in column A (index 1).
    """
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        cell_value = ws.cell(row=row_idx, column=1).value
        # Handle both integer and string comparisons
        if cell_value is not None:
            try:
                if int(cell_value) == int(group_id):
                    return row_idx
            except (ValueError, TypeError):
                continue
    return None

def update_timestamp_only(ws, row_idx: int):
    """
    Updates only the Timestamp column for an existing row.
    Use this function when a row has already been analyzed but you want to refresh its check time.
    """
    ts = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    # Timestamp is in column H (index 8)
    ws.cell(row=row_idx, column=8).value = ts

def update_or_append_row(ws, group_id: int, artifact: str, input_hex_name: str, hidden, leak, limit):
    """
    Updates an existing row if GroupID exists, otherwise appends a new row.
    """
    # RugPull is True if any flag is True
    rugpull = bool(hidden) or bool(leak) or bool(limit)
    ts = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    
    # Try to find existing row
    row_idx = find_row_by_group_id(ws, group_id)
    
    if row_idx is not None:
        # Update existing row
        print(f"  [UPDATE] Updating existing row {row_idx} for GroupID {group_id}")
        ws.cell(row=row_idx, column=1).value = group_id
        ws.cell(row=row_idx, column=2).value = artifact
        ws.cell(row=row_idx, column=3).value = input_hex_name
        ws.cell(row=row_idx, column=4).value = hidden
        ws.cell(row=row_idx, column=5).value = leak
        ws.cell(row=row_idx, column=6).value = limit
        ws.cell(row=row_idx, column=7).value = rugpull
        ws.cell(row=row_idx, column=8).value = ts
    else:
        # Append new row
        print(f"  [NEW] Appending new row for GroupID {group_id}")
        ws.append([group_id, artifact, input_hex_name, hidden, leak, limit, rugpull, ts])

def write_log(group_id: int, tag: str, stdout: str, stderr: str):
    """Writes the analysis output to a log file."""
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    log_file = LOG_DIR / f"{group_id}-{tag}-{ts}.log"
    with log_file.open("w", encoding="utf-8") as f:
        f.write("=== STDOUT ===\n")
        f.write(stdout or "")
        f.write("\n\n=== STDERR ===\n")
        f.write(stderr or "")
    return log_file

# ----------------------------
# Main
# ----------------------------
def main():
    ensure_dirs()
    wb, ws = init_workbook()

    # Iterate group folders: bytecode/2, /3, ..., /1001
    for group_id in range(2, 1002):
        row_idx = find_row_by_group_id(ws, group_id)
        artifact_status = None
        
        # 1. Check the status of the existing row
        if row_idx is not None:
            # Artifact is in column B (index 2)
            artifact_status = ws.cell(row=row_idx, column=2).value
            
            # A. If it was previously analyzed successfully (Artifact points to a real hex file), leave the row untouched
            if artifact_status is not None and artifact_status not in ["(missing group dir)", "no hex file"]:
                print(f"[SKIP] GroupID {group_id} successfully analyzed before ({artifact_status}). Skipping all.")
                continue # Skip to the next GroupID

            # B. If it previously had a defect status (missing dir or no hex file), proceed to step 2 to re-check its current state.
        
        # 2. Check if the group directory exists
        group_dir = COMPILED_DIR / str(group_id)
        
        if not group_dir.exists():
            current_status = "(missing group dir)"
            
            # If it still has the same defect status as before, only update the Timestamp.
            if row_idx is not None and artifact_status == current_status:
                update_timestamp_only(ws, row_idx)
                save_workbook(wb)
                print(f"[WARN] Group folder {group_dir} still missing. Updating timestamp only.")
                continue
            
            # If the row is new or its status has changed, add/update the row accordingly.
            update_or_append_row(ws, group_id, current_status, "", None, None, None)
            save_workbook(wb)
            print(f"[WARN] Missing group folder: {group_dir}")
            continue

        # 3. Check for hex files (only if group_dir exists)
        input_hex_files = sorted(group_dir.glob("*.hex"))

        if not input_hex_files:
            current_status = "no hex file"
            
            # If it still has the same defect status as before, only update the Timestamp.
            if row_idx is not None and artifact_status == current_status:
                update_timestamp_only(ws, row_idx)
                save_workbook(wb)
                print(f"[WARN] GroupID {group_id} still has no *.hex files. Updating timestamp only.")
                continue
            
            # If the row is new or its status has changed, add/update the row accordingly.
            update_or_append_row(ws, group_id, current_status, "(N/A)", None, None, None)
            save_workbook(wb)
            print(f"[WARN] No *.hex files found in {group_dir}")
            continue

        # 4. Analysis phase (only if the row is new or previously had a defect but now has valid hex files)
        
        # If we reached this point, the row is either new or previously defective, and a valid hex file has now been found.
        for input_hex_path in input_hex_files:
            artifact_name = input_hex_path.name

            print(f"\n=== Analyzing {artifact_name} (in group {group_id}) ===")

            # Clear cache using the hex file's name
            clear_contract_cache(input_hex_path)

            # Run analysis directly on the .hex file
            ok, out, err = run_crpwarner_on(input_hex_path)
            write_log(group_id, input_hex_path.stem, out, err)

            if not ok:
                # If analysis fails but the hex file exists, update the row with the real Artifact path.
                update_or_append_row(ws, group_id, artifact_name, artifact_name, None, None, None)
                save_workbook(wb)
                print(f"[ERROR] Analysis failed for {artifact_name}. See logs.")
                continue

            hidden, leak, limit = parse_flags(out)
            update_or_append_row(ws, group_id, artifact_name, artifact_name, hidden, leak, limit)
            save_workbook(wb)

            time.sleep(0.15)  # keeps logs tidy

    print(f"\nAll done. Results: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()